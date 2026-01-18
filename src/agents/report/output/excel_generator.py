"""Excel generation for business intelligence reports.

Uses openpyxl for Excel file generation with multiple worksheets.
"""

import asyncio
import io
import logging
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, LineChart, Reference

from ..schemas import Report, ReportType

logger = logging.getLogger(__name__)


async def generate_excel(report: Report) -> bytes:
    """Generate Excel report from Report object.

    Args:
        report: Report object with all data

    Returns:
        Excel file as bytes
    """
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, lambda: _generate_excel_sync(report))


def _generate_excel_sync(report: Report) -> bytes:
    """Generate Excel file synchronously."""
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create Summary sheet
    _create_summary_sheet(wb, report)

    # Create data sheets based on report type
    if report.report_type == ReportType.EXPENSE_SUMMARY and report.expense_data:
        _create_expense_sheets(wb, report)
    elif report.report_type == ReportType.VENDOR_ANALYSIS and report.vendor_data:
        _create_vendor_sheets(wb, report)
    elif report.report_type == ReportType.INVOICE_RECONCILIATION and report.reconciliation_data:
        _create_reconciliation_sheets(wb, report)

    # Create Insights sheet
    if report.insights:
        _create_insights_sheet(wb, report)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _apply_header_style(cell):
    """Apply header styling to a cell."""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_currency_format(cell):
    """Apply currency format to a cell."""
    cell.number_format = '"$"#,##0.00'


def _apply_percent_format(cell):
    """Apply percentage format to a cell."""
    cell.number_format = '0.0%'


def _create_summary_sheet(wb: Workbook, report: Report):
    """Create the summary sheet."""
    ws = wb.create_sheet("Summary")

    # Title
    ws['A1'] = f"{report.report_type.value.replace('_', ' ').title()} Report"
    ws['A1'].font = Font(bold=True, size=18, color="1a73e8")
    ws.merge_cells('A1:D1')

    # Report details
    ws['A3'] = "Report ID:"
    ws['B3'] = report.id
    ws['A4'] = "Generated:"
    ws['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['A5'] = "Date Range:"
    ws['B5'] = f"{report.date_range.start} to {report.date_range.end}"

    # Summary metrics
    if report.summary:
        ws['A7'] = "Summary Metrics"
        ws['A7'].font = Font(bold=True, size=14)

        metrics = [
            ("Total Amount", report.summary.total_amount, True),
            ("Documents", report.summary.document_count, False),
            ("Records", report.summary.record_count, False),
            ("Vendors", report.summary.vendor_count, False),
            ("Categories", report.summary.category_count, False),
            ("Top Category", report.summary.top_category or "N/A", False),
            ("Top Vendor", report.summary.top_vendor or "N/A", False),
            ("Avg Transaction", report.summary.avg_transaction_amount, True),
        ]

        for i, (label, value, is_currency) in enumerate(metrics, start=9):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'] = value
            if is_currency and isinstance(value, (int, float)):
                _apply_currency_format(ws[f'B{i}'])

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30


def _create_expense_sheets(wb: Workbook, report: Report):
    """Create expense report specific sheets."""
    data = report.expense_data

    # Category breakdown sheet
    ws = wb.create_sheet("By Category")

    headers = ["Category", "Amount", "% of Total", "Transactions"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)

    for row, item in enumerate(data.by_category, start=2):
        ws.cell(row=row, column=1, value=item.category)
        cell = ws.cell(row=row, column=2, value=item.amount)
        _apply_currency_format(cell)
        cell = ws.cell(row=row, column=3, value=item.percentage / 100)
        _apply_percent_format(cell)
        ws.cell(row=row, column=4, value=item.count)

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15

    # Vendor breakdown sheet
    ws = wb.create_sheet("By Vendor")

    headers = ["Vendor", "Total Spend", "% of Total", "Invoices", "Avg per Invoice"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)

    for row, item in enumerate(data.by_vendor, start=2):
        ws.cell(row=row, column=1, value=item.vendor)
        cell = ws.cell(row=row, column=2, value=item.amount)
        _apply_currency_format(cell)
        cell = ws.cell(row=row, column=3, value=item.percentage / 100)
        _apply_percent_format(cell)
        ws.cell(row=row, column=4, value=item.invoice_count)
        cell = ws.cell(row=row, column=5, value=item.avg_amount)
        _apply_currency_format(cell)

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15

    # Monthly trends sheet
    if data.monthly_trends:
        ws = wb.create_sheet("Monthly Trends")

        headers = ["Month", "Amount", "Transactions"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            _apply_header_style(cell)

        for row, item in enumerate(data.monthly_trends, start=2):
            ws.cell(row=row, column=1, value=item.month)
            cell = ws.cell(row=row, column=2, value=item.amount)
            _apply_currency_format(cell)
            ws.cell(row=row, column=3, value=item.count)

        # Add line chart
        if len(data.monthly_trends) >= 2:
            chart = LineChart()
            chart.title = "Monthly Spending Trend"
            chart.x_axis.title = "Month"
            chart.y_axis.title = "Amount ($)"

            data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(data.monthly_trends) + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(data.monthly_trends) + 1)

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            chart.width = 15
            chart.height = 8

            ws.add_chart(chart, "E2")

    # Top expenses sheet
    if data.top_expenses:
        ws = wb.create_sheet("Top Expenses")

        # Use dynamic headers based on data
        if data.top_expenses:
            headers = list(data.top_expenses[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
                _apply_header_style(cell)

            for row, item in enumerate(data.top_expenses, start=2):
                for col, header in enumerate(headers, 1):
                    value = item.get(header, "")
                    ws.cell(row=row, column=col, value=value)


def _create_vendor_sheets(wb: Workbook, report: Report):
    """Create vendor analysis specific sheets."""
    data = report.vendor_data

    # Vendor rankings sheet
    ws = wb.create_sheet("Vendor Rankings")

    headers = ["Rank", "Vendor", "Total Spend", "% of Total", "Invoices", "Avg per Invoice"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)

    for row, item in enumerate(data.vendor_rankings, start=2):
        ws.cell(row=row, column=1, value=row - 1)  # Rank
        ws.cell(row=row, column=2, value=item.vendor)
        cell = ws.cell(row=row, column=3, value=item.amount)
        _apply_currency_format(cell)
        cell = ws.cell(row=row, column=4, value=item.percentage / 100)
        _apply_percent_format(cell)
        ws.cell(row=row, column=5, value=item.invoice_count)
        cell = ws.cell(row=row, column=6, value=item.avg_amount)
        _apply_currency_format(cell)

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15

    # Add bar chart
    if len(data.vendor_rankings) >= 2:
        chart = BarChart()
        chart.title = "Top Vendors by Spend"
        chart.type = "bar"
        chart.style = 10

        data_ref = Reference(ws, min_col=3, min_row=1, max_row=min(len(data.vendor_rankings), 10) + 1)
        cats = Reference(ws, min_col=2, min_row=2, max_row=min(len(data.vendor_rankings), 10) + 1)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.width = 15
        chart.height = 10

        ws.add_chart(chart, "H2")

    # Concentration metrics sheet
    ws = wb.create_sheet("Concentration")

    metrics = data.concentration_metrics or {}
    ws['A1'] = "Vendor Concentration Metrics"
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = "Top 3 Vendors (% of Total)"
    cell = ws['B3'] = metrics.get("top_3_percentage", 0) / 100
    _apply_percent_format(cell)

    ws['A4'] = "Total Vendors"
    ws['B4'] = metrics.get("vendor_count", 0)

    ws['A5'] = "Average per Vendor"
    cell = ws['B5'] = metrics.get("avg_per_vendor", 0)
    _apply_currency_format(cell)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def _create_reconciliation_sheets(wb: Workbook, report: Report):
    """Create reconciliation specific sheets."""
    data = report.reconciliation_data

    # Summary sheet already exists, add reconciliation details
    ws = wb.create_sheet("Reconciliation Summary")

    summary = data.summary or {}

    ws['A1'] = "Reconciliation Summary"
    ws['A1'].font = Font(bold=True, size=14)

    metrics = [
        ("Total Records", summary.get("total_records", 0)),
        ("Matched Records", summary.get("matched_count", 0)),
        ("Missing PO References", summary.get("missing_po_count", 0)),
    ]

    for i, (label, value) in enumerate(metrics, start=3):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'] = value

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15


def _create_insights_sheet(wb: Workbook, report: Report):
    """Create insights sheet."""
    ws = wb.create_sheet("Insights")

    headers = ["#", "Category", "Title", "Description", "Severity"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)

    for row, insight in enumerate(report.insights, start=2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=insight.category)
        ws.cell(row=row, column=3, value=insight.title)
        ws.cell(row=row, column=4, value=insight.description)
        cell = ws.cell(row=row, column=5, value=insight.severity or "info")

        # Color-code severity
        if insight.severity == "warning":
            cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
        elif insight.severity == "critical":
            cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")

    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 12

    # Set row heights and wrap text for description
    for row in range(2, len(report.insights) + 2):
        ws.row_dimensions[row].height = 45
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
